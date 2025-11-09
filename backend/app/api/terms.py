"""API routes for interacting with legal terms."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from docx import Document
from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook

from ..models.terms import Term, TermMergeResult, TermsRepository

router = APIRouter(prefix="/terms", tags=["terms"])

_DATA_PATH = Path(__file__).resolve().parents[2] / "data" / "terms.json"
_repository = TermsRepository(_DATA_PATH)


def _parse_xlsx(data: bytes) -> list[Term]:
    workbook = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise ValueError("Workbook is empty") from exc

    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        if not isinstance(value, str):
            continue
        normalized = value.strip().lower()
        if normalized in {"zh", "bn", "en"} and normalized not in header_map:
            header_map[normalized] = index

    missing = {label for label in ("zh", "bn", "en") if label not in header_map}
    if missing:
        raise ValueError("Workbook must contain zh, bn, and en columns")

    terms: list[Term] = []
    for row in rows:
        if row is None:
            continue
        zh = row[header_map["zh"]] if header_map["zh"] < len(row) else None
        bn = row[header_map["bn"]] if header_map["bn"] < len(row) else None
        en = row[header_map["en"]] if header_map["en"] < len(row) else None

        if not any([zh, bn, en]):
            continue

        term = Term(zh=str(zh or "").strip(), bn=str(bn or "").strip(), en=str(en or "").strip())
        if not (term.zh and term.bn and term.en):
            # Skip incomplete rows to avoid introducing partial entries.
            continue
        terms.append(term)

    if not terms:
        raise ValueError("No valid terms found in the provided workbook")

    return terms


def _parse_docx(data: bytes) -> list[Term]:
    document = Document(BytesIO(data))
    terms: list[Term] = []

    for paragraph in document.paragraphs:
        text = paragraph.text.strip()
        if not text or "｜" not in text:
            continue
        parts = [part.strip() for part in text.split("｜")]
        if len(parts) != 3 or not all(parts):
            continue
        terms.append(Term(zh=parts[0], bn=parts[1], en=parts[2]))

    if not terms:
        raise ValueError("No valid terms found in the provided document")

    return terms


def _import_terms_from_file(upload: UploadFile) -> TermMergeResult:
    suffix = Path(upload.filename or "").suffix.lower()
    data = upload.file.read()
    if not data:
        raise ValueError("Uploaded file is empty")

    if suffix == ".xlsx":
        terms = _parse_xlsx(data)
    elif suffix == ".docx":
        terms = _parse_docx(data)
    else:
        raise ValueError("Unsupported file type. Please upload a .xlsx or .docx file.")

    return _repository.merge_terms(terms)


@router.get("", response_model=list[Term], summary="Search terms")
async def search_terms(q: str | None = Query(default=None, description="Keyword to search for")) -> list[Term]:
    """Perform a fuzzy search over the stored terms."""

    try:
        return _repository.search(q.strip() if q else None)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=str(exc),
        ) from exc


@router.post("/upload", summary="Bulk import terms")
async def upload_terms(file: UploadFile = File(...)) -> dict[str, int]:
    """Upload a spreadsheet or document to import terms in bulk."""

    try:
        merge_result = _import_terms_from_file(file)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
    except Exception as exc:  # pragma: no cover - defensive catch for unexpected parsing errors
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to process the uploaded file.",
        ) from exc

    return {"added": merge_result.added, "total": merge_result.total}


__all__ = ["router"]
